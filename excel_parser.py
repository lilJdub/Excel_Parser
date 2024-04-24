import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog

# 選擇文件=得到路徑
file_path = filedialog.askopenfilename(filetypes=[('Excel Files', ('.xlsx', '.xls'))])
#得到workbook以及現在工作表
workbook = load_workbook(file_path)
sheet = workbook.active

#首列先記下來(之後會需要丟回每一張sheet上面，不然看不到每列要幹嘛)
header=next(sheet.iter_rows(min_row=1, max_row=1))

#從整張表上到下遍歷每一個row
S_row_numbers = []
for row in range(1, sheet.max_row + 1):
    #找每個row的第一格(類型)
    cell_value = sheet.cell(row=row, column=1).value
    #若Value為S則暫存，以這些值切開整張表
    if cell_value == 'S':
        S_row_numbers.append(row)
#add 最後一行(為了定錨sheet尾巴)
S_row_numbers.append(sheet.max_row)

# 確定有一個表以上(其實好像不寫這行也ok)
if len(S_row_numbers) > 1:
    #新的excel檔workbook
    new_workbook = Workbook()
    
    #從第一個S開始遍歷到倒數第二個S(得到每一個sheet的表頭)(也就是這個迴圈跑一次就是一張表)
    for indx in range(0, len(S_row_numbers) - 1):
        #各顏色filling先暫存
        color_greenfill=[]
        color_mintfill=[]
        color_orangefill=[0]
        
        # 創建新的工作表=>獲取sheet name(每個表頭那列的第三格)
        sheet_name=sheet.cell(row=S_row_numbers[indx], column=3).value
        #解決亂碼
        sheet_name = re.sub(r'[\\/*?[\]:]', '', str(sheet_name))

        #做出sheet:加入名字
        print("正在處理: "+str(sheet_name))
        new_sheet = new_workbook.create_sheet(title=sheet_name)
        #把剛剛最上面的header加回去
        new_sheet.append([cell.value for cell in header])

        #Sheet的開頭= 紀錄的"S"的那行
        start_row = S_row_numbers[indx]
        #Sheet的尾巴= 下一個表頭的上一行
        end_row = S_row_numbers[indx + 1] - 1
        #開始複製整張表的數據，整張表挑出sheet的開頭和尾巴，中間一行一行看: enumerate會同時可以操作值以及序列順序
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=sheet.max_column)):
            #每一row的數據貼進新sheet裡
            new_sheet.append([cell.value for cell in row])
            
            """到這邊其實值都處理好了，現在來填顏色"""
            #獲得該填綠色，薄荷和橘色的行數，因為enumerate會從0開始起算要記得+1才能拿到
            if row[0].value=="S":
                color_greenfill.append(row_idx+1)
            elif row[0].value=="SC":
                color_mintfill.append(row_idx+1)
            elif row[0].value=="STEP":
                color_orangefill.append(row_idx+1)

        #獲取顏色之後在新表內填入綠色
        for row_index in color_greenfill:
            # 按照索引取得該行
            row = new_sheet[row_index+1]
            # 將該行每個單元格的背景色設置為綠色
            for cell in row:
                cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

        #在新表內填入薄荷綠色
        for row_index in color_mintfill:
            # 按照索引取得該行
            row = new_sheet[row_index+1]
            # 將該行每個單元格的背景色設置為薄荷綠色
            for cell in row:
                cell.fill = PatternFill(start_color="A8B782", end_color="A8B782", fill_type="solid")
        
        #在新表內填入橘色
        for row_index in color_orangefill:
            # 按照索引取得該行
            row = new_sheet[row_index+1]
            # 將該行每個單元格的背景色設置為橘色
            for cell in row:
                cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        
        #將rows去做分組:因為用SC來做group 我這邊就拿剛剛存好的color_mintfill去作手腳
        group_rows = color_mintfill
        # 添加最後一行的索引(grouping底部)
        group_rows.append(new_sheet.max_row + 1)  
        #遍歷每一個表頭(最後一個是底部所以不要跑到)
        for i in range(0, len(group_rows) - 1):
            #group表下兩行開始折疊(也就是空一行的意思)
            new_sheet.row_dimensions.group(group_rows[i]+2, group_rows[i + 1])
            
    # 保存新的工作簿
    new_file_path = file_path.replace('.xlsx', '_extracted.xlsx')  # 新文件名
    new_workbook.save(new_file_path)
    print("切片完成:文件保存為", new_file_path)
else:
    print("未找到匹配的行數")